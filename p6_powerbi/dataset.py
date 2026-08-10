"""DB -> star-schema Excel workbook that Power BI connects to (live).

Read path only: reads what the metrics engine already stored per snapshot and
reshapes it into fact/dimension tables Power BI likes. Never recomputes a
number. Best-effort by design — callers (the import hook) wrap in try/except so
a dataset write can never break an import.

Tables (each written to its own sheet as a named Excel Table so Power Query can
reference it by name):
  dim_project    project master (for slicers)
  dim_snapshot   one row per import (the period/date dimension)
  fact_metrics   EVM numbers per snapshot (denormalised with project + date)
  fact_category  per-category planned/actual per snapshot
"""
import os
from datetime import datetime, date

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import db
from p6_powerbi import paths


def _naive(dt):
    """Excel/openpyxl reject tz-aware datetimes — drop the tzinfo."""
    return dt.replace(tzinfo=None) if getattr(dt, 'tzinfo', None) else dt


def _parse_date(value):
    """Best-effort ISO string -> naive date/datetime so Power BI gets a real date type."""
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return _naive(value)
    if isinstance(value, date):
        return value
    try:
        return _naive(datetime.fromisoformat(str(value)))
    except ValueError:
        try:
            return datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
        except ValueError:
            return None  # never emit a raw string into a dateTime column (breaks refresh)


def _build_tables(conn):
    """Return {table_name: [header_tuple, *data_tuples]} from the DB."""
    dim_project = [('project_id', 'p6_project_id', 'project_name')]
    for p in conn.execute('SELECT id, p6_project_id, name FROM projects ORDER BY id'):
        dim_project.append((p['id'], p['p6_project_id'] or '', p['name'] or ''))

    dim_snapshot = [('snapshot_id', 'project_id', 'project_name',
                     'data_date', 'imported_at', 'activity_count')]
    fact_metrics = [('snapshot_id', 'project_id', 'project_name', 'data_date',
                     'pv', 'ev', 'ac', 'spi', 'cpi', 'delay_days',
                     'overall_planned_pct', 'overall_actual_pct', 'variance')]
    snap_rows = conn.execute(
        '''SELECT s.id AS snapshot_id, s.project_id, p.name AS project_name,
                  s.data_date, s.imported_at, s.activity_count,
                  m.pv, m.ev, m.ac, m.spi, m.cpi, m.delay_days,
                  m.overall_planned_pct, m.overall_actual_pct, m.variance
           FROM snapshots s
           JOIN projects p ON p.id = s.project_id
           LEFT JOIN metrics m ON m.snapshot_id = s.id
           ORDER BY s.project_id, s.data_date, s.id'''
    ).fetchall()
    for s in snap_rows:
        dd = _parse_date(s['data_date'])
        dim_snapshot.append((s['snapshot_id'], s['project_id'], s['project_name'],
                             dd, _parse_date(s['imported_at']), s['activity_count']))
        fact_metrics.append((
            s['snapshot_id'], s['project_id'], s['project_name'], dd,
            s['pv'], s['ev'], s['ac'], s['spi'], s['cpi'], s['delay_days'],
            s['overall_planned_pct'], s['overall_actual_pct'], s['variance'],
        ))

    fact_category = [('snapshot_id', 'project_id', 'project_name', 'data_date',
                      'category', 'weight', 'planned_pct', 'actual_pct',
                      'bac', 'ac', 'activity_count')]
    for c in conn.execute(
        '''SELECT cm.snapshot_id, s.project_id, p.name AS project_name, s.data_date,
                  cm.name AS category, cm.weight, cm.planned_pct, cm.actual_pct,
                  cm.bac, cm.ac, cm.activity_count
           FROM category_metrics cm
           JOIN snapshots s ON s.id = cm.snapshot_id
           JOIN projects p ON p.id = s.project_id
           ORDER BY cm.snapshot_id, cm.id'''
    ).fetchall():
        fact_category.append((
            c['snapshot_id'], c['project_id'], c['project_name'], _parse_date(c['data_date']),
            c['category'], c['weight'], c['planned_pct'], c['actual_pct'],
            c['bac'], c['ac'], c['activity_count'],
        ))

    return {
        'dim_project': dim_project,
        'dim_snapshot': dim_snapshot,
        'fact_metrics': fact_metrics,
        'fact_category': fact_category,
    }


def _add_named_table(ws, name, n_cols, n_data_rows):
    """Register the sheet's data as a named Excel Table so Power Query can bind
    to it by name. A header-only table (0 data rows, empty DB) is valid — no
    padding, so Power BI never sees a phantom blank row."""
    ref = f'A1:{get_column_letter(n_cols)}{n_data_rows + 1}'
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name='TableStyleLight1', showRowStripes=False)
    ws.add_table(table)


def write_dataset(workbook_path=None, conn=None):
    """Write the star-schema workbook Power BI reads.

    Returns {'workbook': path, 'tables': {name: data_row_count}}.
    """
    close = False
    if conn is None:
        conn = db.get_conn()
        close = True
    try:
        tables = _build_tables(conn)
    finally:
        if close:
            conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    counts = {}
    for name, rows in tables.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(list(row))
        counts[name] = len(rows) - 1
        _add_named_table(ws, name, n_cols=len(rows[0]), n_data_rows=len(rows) - 1)

    path = workbook_path or paths.dataset_workbook()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    # Write to a temp file then atomically replace, so a concurrent Power BI
    # refresh never reads a half-written workbook. If the target is locked
    # (Power BI has it open), keep the last good copy rather than crashing.
    tmp = path + '.tmp'
    wb.save(tmp)
    try:
        os.replace(tmp, path)
    except OSError:
        try:
            os.remove(tmp)
        except OSError:
            pass
    return {'workbook': path, 'tables': counts}
