"""Excel workflow — KB-driven template, parse-back, and the Productivity Mapping Sheet
(including the Knowledge-Not-Available status surviving into the export)."""
import os

from openpyxl import Workbook, load_workbook

from p6_prodkb import excelio, kb
from p6_prodkb import match as m

BASE = os.path.join(os.path.dirname(__file__), "..", "productivity_kb")


def _t():
    return kb.load_templates(bundled=BASE, overlay=[])


def test_classifications_from_kb():
    cl = excelio.classifications(_t())
    assert "Residential/Villa" in cl["project_types"]
    assert "Industrial" in cl["project_types"]
    assert "GLOBAL" not in cl["project_types"]          # neutral fallback is not a "choice"
    assert any("column" in w.lower() for w in cl["work_types"])
    assert "m2" in cl["units"] and "m3" in cl["units"]


def test_template_is_kb_driven(tmp_path):
    p = str(tmp_path / "template.xlsx")
    excelio.write_template(p, _t())
    wb = load_workbook(p)
    assert [c.value for c in wb["Activities"][1]] == excelio.INPUT_HEADERS
    assert "KB Reference" in wb.sheetnames
    ref_ptypes = [r[0] for r in wb["KB Reference"].iter_rows(min_row=2, values_only=True)]
    assert "Residential/Villa" in ref_ptypes            # reference lists come from the live KB


def test_read_activities(tmp_path):
    p = str(tmp_path / "filled.xlsx")
    wb = Workbook(); ws = wb.active; ws.title = "Activities"
    ws.append(excelio.INPUT_HEADERS)
    ws.append(["A-100", "Internal Wall Plaster", "Residential/Villa", "", 300, "m2", ""])
    ws.append([None, None, None, None, None, None, None])   # blank row skipped
    wb.save(p)
    acts = excelio.read_activities(p)
    assert len(acts) == 1
    assert acts[0]["name"] == "Internal Wall Plaster"
    assert acts[0]["project_type"] == "Residential/Villa"
    assert acts[0]["quantity"] == 300.0


def test_batch_resolve_and_mapping_export(tmp_path):
    tmpls = _t()
    activities = [
        {"activity_id": "A1", "name": "Internal Wall Plaster", "project_type": "Residential/Villa", "quantity": 300, "unit": "m2"},
        {"activity_id": "A2", "name": "Specialized Process Equipment Installation", "project_type": "Oil & Gas", "quantity": 5, "unit": "unit"},
        {"activity_id": "A3", "name": "RC Columns L3", "project_type": "Industrial", "quantity": None, "unit": "m3"},
    ]
    rows = []
    for a in activities:
        r = m.resolve({"name": a["name"]}, a.get("quantity"),
                      project_type=a.get("project_type"), templates=tmpls)
        rows.append({"input": a, "result": r["result"],
                     "status": r.get("status"), "status_label": r.get("status_label")})
    assert [r["status"] for r in rows] == ["ok", "knowledge_not_available", "needs_input"]

    p = str(tmp_path / "mapping.xlsx")
    excelio.write_mapping_sheet(p, rows)
    data = list(load_workbook(p)["Productivity Rates"].iter_rows(values_only=True))
    assert data[0] == tuple(excelio.MAPPING_HEADERS)
    assert data[1][7] == 25 and data[1][14] == "Calculated"          # villa rate exported
    assert data[2][14] == "Knowledge Not Available"                   # KB gap survives export
    assert data[2][7] in ("", None)                                   # and carries NO invented rate
