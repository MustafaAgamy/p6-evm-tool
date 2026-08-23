"""Critical Path Analyzer — PDF (HTML) and Excel exporters."""
from datetime import datetime
import report_theme
from p6_evm.parser import ScheduleData
from p6_critpath.analysis import build_report
from p6_critpath.exporters import render_html, to_excel


def _sched(data_date, chain, ms_finish, ms_bl, ms_tf, extra):
    d = ScheduleData()
    d.project = {'name': 'T', 'data_date': data_date}
    d.wbs = {}; d.activities = {}; d.relationships = []; d.baseline_by_id = {}
    prev = None
    for i, (w, tf) in enumerate(chain):
        wid = f'W_{w}'; d.wbs[wid] = {'name': w, 'parent_object_id': None}; oid = f'A{i}'
        fin = datetime(2026, 8, 1 + i)
        d.activities[oid] = {'id': f'ACT{i}', 'name': w, 'task_type': 'Task', 'calendar_id': None,
                             'wbs_id': wid, 'percent_complete': 0.0, 'total_float_days': tf,
                             'remaining_early_finish': fin, 'planned_finish': fin, 'object_id': oid}
        d.baseline_by_id[f'ACT{i}'] = {'planned_start': fin, 'planned_finish': fin}
        if prev is not None:
            d.activities[prev]['remaining_early_finish'] = fin
            d.relationships.append({'pred_id': prev, 'succ_id': oid, 'type': 'FS', 'lag_days': 0.0})
        prev = oid
    d.activities['MS'] = {'id': 'M999', 'name': 'Project Completion', 'task_type': 'FinishMilestone',
                          'calendar_id': None, 'wbs_id': None, 'total_float_days': ms_tf,
                          'remaining_early_finish': ms_finish, 'planned_finish': ms_finish}
    d.activities[prev]['remaining_early_finish'] = ms_finish
    d.relationships.append({'pred_id': prev, 'succ_id': 'MS', 'type': 'FS', 'lag_days': 0.0})
    d.baseline_by_id['M999'] = {'planned_start': data_date, 'planned_finish': ms_bl}
    for i, tf in enumerate(extra):
        d.activities[f'X{i}'] = {'id': f'X{i}', 'name': f'x{i}', 'task_type': 'Task', 'calendar_id': None,
                                 'total_float_days': tf, 'object_id': f'X{i}'}
    return d


def _report():
    prev = _sched(datetime(2026, 6, 30), [('Foundations', 0.0), ('Structure', 0.0), ('Roof', 0.0)],
                  datetime(2027, 1, 5), datetime(2026, 12, 10), -5.0, [5.0, 20.0])
    curr = _sched(datetime(2026, 7, 19), [('Foundations', 0.0), ('MEP', 0.0), ('Commissioning', 0.0)],
                  datetime(2027, 1, 23), datetime(2026, 12, 10), -20.0, [-1.0, 6.0])
    r = build_report({'previous': prev, 'current': curr}, 'two_updates')
    r['files'] = {'previous': 'prev.xml', 'current': 'curr.xml'}
    return r


def test_render_html_has_all_sections():
    html = render_html(_report())
    for key in ('verdict', 'dashboard', 'driving_path', 'census', 'milestones', 'float_migration', 'recommendation'):
        assert f'data-sec="{key}"' in html
    assert 'NEW ON PATH' in html            # the reroute highlight survives into the PDF
    assert 'Critical Path Analyzer' in html


def test_render_html_section_filter():
    html = render_html(_report(), sections=['census'])
    assert 'data-sec="census"' in html
    assert 'data-sec="dashboard"' not in html


def test_render_html_theme_dark():
    html = render_html(_report(), theme='dark')
    assert 'data-rpt-theme="dark"' in html
    assert report_theme.THEMES['dark']['rpt-accent'] in html  # '#5b9bff'


def test_render_html_theme_default_is_light():
    html = render_html(_report())
    assert html.startswith('<!doctype html>') and html.rstrip().endswith('</html>')
    assert 'data-rpt-theme="light"' in html


def test_to_excel_writes_expected_sheets(tmp_path):
    from openpyxl import load_workbook
    out = tmp_path / 'cpa.xlsx'
    to_excel(_report(), str(out))
    wb = load_workbook(out)
    assert set(wb.sheetnames) >= {'Census', 'Milestones', 'Driving path (current)', 'Float migration'}
    assert wb['Census']['A1'].value == 'Measure'
