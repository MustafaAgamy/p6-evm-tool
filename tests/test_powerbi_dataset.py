"""Tests for p6_powerbi.dataset — DB -> star-schema Excel workbook."""
import os

import openpyxl
import pytest

import db
import p6_powerbi.paths as paths
from p6_powerbi.dataset import write_dataset


def _seed_one(pid_name='Grain Terminal'):
    pid = db.upsert_project('P001', pid_name)
    sid = db.insert_snapshot(pid, '2026-07-31', '/p.xml', '/c.xml', 'h1', 100, 3)
    db.insert_metrics(sid, {
        'pv': 1000, 'ev': 900, 'ac': 950, 'spi': 0.9, 'cpi': 0.95,
        'delay_days': 12, 'overall_planned_pct': 0.6, 'overall_actual_pct': 0.54,
        'variance': -100,
    })
    db.insert_category_metrics(sid, {
        'Construction': {'weight': 0.9, 'planned_pct': 0.6, 'actual_pct': 0.5,
                         'bac': 1000, 'ac': 900, 'activity_count': 50, 'overridden': False},
    })
    return pid, sid


def test_dataset_workbook_path(monkeypatch, tmp_path):
    monkeypatch.setattr(paths, 'app_data_dir', lambda: str(tmp_path))
    wb = paths.dataset_workbook()
    assert wb.endswith(os.path.join('powerbi', 'dataset', 'p6evm.xlsx'))
    assert os.path.isdir(os.path.dirname(wb))


def test_write_dataset_creates_named_tables(temp_db, tmp_path):
    _seed_one()
    out = tmp_path / 'p6evm.xlsx'
    result = write_dataset(workbook_path=str(out))

    assert os.path.exists(out)
    assert result['tables']['fact_metrics'] == 1
    assert result['tables']['fact_category'] == 1

    wb = openpyxl.load_workbook(out)
    assert {'dim_project', 'dim_snapshot', 'fact_metrics', 'fact_category'}.issubset(set(wb.sheetnames))
    # each sheet exposes a named Excel Table Power Query can bind to
    assert 'fact_metrics' in wb['fact_metrics'].tables
    assert 'fact_category' in wb['fact_category'].tables

    ws = wb['fact_metrics']
    headers = [c.value for c in ws[1]]
    assert 'spi' in headers and 'project_name' in headers and 'data_date' in headers
    spi_col = headers.index('spi') + 1
    assert ws.cell(row=2, column=spi_col).value == pytest.approx(0.9)


def test_write_dataset_multiple_snapshots_for_trend(temp_db, tmp_path):
    pid = db.upsert_project('P1', 'Proj')
    for i, d in enumerate(['2026-05-31', '2026-06-30', '2026-07-31']):
        sid = db.insert_snapshot(pid, d, '/p.xml', '/c.xml', f'h{i}', 100, 3)
        db.insert_metrics(sid, {'spi': 0.9 - i * 0.02, 'cpi': 1.0, 'delay_days': i * 5})
    out = tmp_path / 'p6evm.xlsx'
    result = write_dataset(workbook_path=str(out))
    assert result['tables']['fact_metrics'] == 3
    # data_date written as a real date type (not a bare string)
    wb = openpyxl.load_workbook(out)
    ws = wb['fact_metrics']
    headers = [c.value for c in ws[1]]
    dd_col = headers.index('data_date') + 1
    from datetime import date, datetime as _dt
    assert isinstance(ws.cell(row=2, column=dd_col).value, (date, _dt))


def test_parse_date_malformed_returns_none():
    # A bad date must never become a raw string in a dateTime column (breaks Power BI refresh).
    from p6_powerbi.dataset import _parse_date
    assert _parse_date('not-a-date') is None
    assert _parse_date('') is None
    assert _parse_date(None) is None
    from datetime import date
    assert _parse_date('2026-07-31') in (date(2026, 7, 31),) or _parse_date('2026-07-31').year == 2026


def test_write_dataset_empty_db_headers_only(temp_db, tmp_path):
    out = tmp_path / 'p6evm.xlsx'
    result = write_dataset(workbook_path=str(out))
    assert os.path.exists(out)
    assert result['tables']['fact_metrics'] == 0
    # a valid named table still exists (padded empty body row)
    wb = openpyxl.load_workbook(out)
    assert 'fact_metrics' in wb['fact_metrics'].tables
