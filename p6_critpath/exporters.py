"""PDF (print HTML) and Excel exporters for the Critical Path Analyzer.

render_html builds a landscape, print-ready page from the report the client already
holds (no re-parse). Every section is wrapped in <section data-sec="KEY"> so the
preview's Report Contents picker can toggle it and the PDF re-flows. Print colours are
theme tokens (`--rpt-*`, injected via report_theme.theme_style_tag) so the same markup
renders in every appearance mode — see report_theme.py.
"""

import report_theme

# Module colour constants, resolved at render time via the injected `--rpt-*` custom
# properties (see report_theme.py). Verified usage before mapping:
#   _BAD/_WARN/_GOOD/_MUTED -> the census/milestone/dashboard "critical / near / safe"
#                              semantic states (red/amber/green/grey), unchanged meaning
#   _ACCENT                 -> the page's one brand blue, reused everywhere as heading /
#                              link / "current"-role colour, not a chart-series colour
#   _GRID / _AXIS_TXT       -> chart gridlines/reference lines vs. bar/point value labels
#   Driving-path lane states (NEW ON PATH / LEFT PATH / stayed / complete):
#     - complete (bdone)  -> _GOOD  (already green; finished work, unambiguous)
#     - new on path (bnew)-> _BAD   (already red in the source design; a newly-critical
#                                    front is the headline risk of the report)
#     - left path (bleft) -> NEUTRAL (hair-strong/surface-2/muted; kept neutral as in the
#                                    source design — a front leaving the critical path isn't
#                                    inherently bad, and must stay visually distinct from the
#                                    red "new on path" alarm)
#     - stayed (default)  -> neutral edge/surface tokens, no semantic colour
_BAD = report_theme.var('rpt-bad')
_BAD_BG = report_theme.var('rpt-bad-bg')
_WARN = report_theme.var('rpt-warn')
_GOOD = report_theme.var('rpt-good')
_MUTED = report_theme.var('rpt-muted')
_ACCENT = report_theme.var('rpt-accent')
_GRID = report_theme.var('rpt-chart-grid')
_AXIS_TXT = report_theme.var('rpt-chart-axis')
_EDGE = report_theme.var('rpt-edge')
_BG = report_theme.var('rpt-bg')

_ROLE_LABEL = {'baseline': 'Baseline', 'previous': 'Previous update', 'current': 'Current update'}
_ROLE_SHORT = {'baseline': 'Baseline', 'previous': 'Previous', 'current': 'Current'}
_STATUS_HEX = {'good': (_GOOD, report_theme.var('rpt-good-bg')), 'warn': (_WARN, report_theme.var('rpt-warn-bg')),
               'bad': (_BAD, _BAD_BG)}


def _e(v):
    if v is None:
        return ''
    return (str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def _sd(v, suffix=' d'):
    return '—' if v is None else f"{'+' if v > 0 else ''}{v}{suffix}"


def _cpli(v):
    return 'n/a' if v is None else f"{v:.2f}"


# ── Sections ─────────────────────────────────────────────────────────────────

def _header(report):
    files = report.get('files', {})
    parts = []
    for role in ('baseline', 'previous', 'current'):
        if role in files:
            parts.append(f"{_ROLE_LABEL[role]}: <b>{_e(files[role])}</b>")
    # Data-date window strip — one pill per loaded schedule.
    dd = report.get('data_dates', {})
    strip = []
    for role in ('baseline', 'previous', 'current'):
        if dd.get(role):
            strip.append(f'<span class="ddpill">{_ROLE_SHORT[role]} data date <b>{_e(dd[role])}</b></span>')
    pills = f'<div class="ddstrip">{"".join(strip)}</div>' if strip else ''
    return (f'<div class="rh"><div><h1>Critical Path Analyzer</h1>'
            f'<div class="meta">{" &nbsp;·&nbsp; ".join(parts)}</div>{pills}</div></div>')


def _banner(report):
    d = report.get('dashboard', {})
    status = d.get('status', 'warn')
    hexc, bg = _STATUS_HEX.get(status, _STATUS_HEX['warn'])
    return (f'<div class="banner" style="background:{bg};border-color:{hexc}">'
            f'<span class="btag" style="background:{hexc}">{_e(d.get("status_label"))}</span>'
            f'<span class="btext">{_e(report.get("conclusion") or d.get("verdict"))}</span></div>')


# ── Dashboard charts (inline SVG / pure-HTML, theme-token colours) ────────────

def _chart_crit_near(cn):
    """Grouped bars: critical (red) + near-critical (amber) per schedule. A null near
    value (a float-less baseline) is labelled 'n/a' rather than drawn as a measured 0."""
    roles = cn.get('roles', [])
    crit = cn.get('critical', [])
    near = cn.get('near', [])
    if not roles:
        return '<div class="note">No census data to plot.</div>'
    vals = [v for v in (list(crit) + list(near)) if v is not None]
    mx = max(vals) if vals else 0
    if mx <= 0:
        mx = 1
    W, H = 300, 170
    pad_l, pad_r, pad_t, pad_b = 8, 8, 22, 34
    plot_h = H - pad_t - pad_b
    plot_w = W - pad_l - pad_r
    n = len(roles)
    group_w = plot_w / n
    bw = min(26.0, group_w / 3)
    gap = 4.0
    y0 = pad_t + plot_h
    parts = [f'<line x1="{pad_l}" y1="{y0}" x2="{W - pad_r}" y2="{y0}" stroke="{_GRID}" stroke-width="1"/>']
    for i, role in enumerate(roles):
        gx = pad_l + i * group_w + group_w / 2
        pairs = [(crit[i] if i < len(crit) else None, _BAD),
                 (near[i] if i < len(near) else None, _WARN)]
        bx0 = gx - (bw * 2 + gap) / 2
        for j, (val, col) in enumerate(pairs):
            x = bx0 + j * (bw + gap)
            if val is None:
                parts.append(f'<text x="{x + bw / 2:.1f}" y="{y0 - 4:.1f}" font-size="8" '
                             f'fill="{_MUTED}" text-anchor="middle">n/a</text>')
                continue
            bh = (val / mx) * plot_h
            y = y0 - bh
            parts.append(f'<rect x="{x:.1f}" y="{y:.1f}" width="{bw:.1f}" height="{bh:.1f}" '
                         f'fill="{col}" rx="2"/>')
            parts.append(f'<text x="{x + bw / 2:.1f}" y="{y - 3:.1f}" font-size="9" '
                         f'font-weight="700" fill="{_AXIS_TXT}" text-anchor="middle">{val}</text>')
        parts.append(f'<text x="{gx:.1f}" y="{H - pad_b + 14:.1f}" font-size="9" '
                     f'fill="{_MUTED}" text-anchor="middle">{_e(_ROLE_SHORT.get(role, role))}</text>')
    return (f'<svg viewBox="0 0 {W} {H}" style="width:100%;height:auto" '
            f'xmlns="http://www.w3.org/2000/svg">{"".join(parts)}</svg>')


def _chart_cpli(ct):
    """CPLI trend polyline across schedules, with dashed 0.95 / 1.00 reference lines.
    Null points are skipped; the ends are inset so the first/last labels aren't clipped."""
    roles = ct.get('roles', [])
    values = ct.get('values', [])
    if not roles:
        return '<div class="note">No CPLI trend to plot.</div>'
    pts = [(i, v) for i, v in enumerate(values) if v is not None]
    W, H = 300, 170
    inset = 30.0
    pad_t, pad_b = 18, 30
    plot_h = H - pad_t - pad_b
    n = len(roles)

    def xpos(i):
        if n <= 1:
            return W / 2
        return inset + (W - 2 * inset) * i / (n - 1)

    allv = [v for _, v in pts] + [0.95, 1.0]
    lo, hi = min(allv), max(allv)
    if hi - lo < 0.1:
        lo -= 0.05
        hi += 0.05
    span = hi - lo

    def ypos(v):
        return pad_t + plot_h * (1 - (v - lo) / span)

    parts = []
    for ref, lbl in ((1.0, '1.00'), (0.95, '0.95')):
        y = ypos(ref)
        parts.append(f'<line x1="{inset}" y1="{y:.1f}" x2="{W - inset}" y2="{y:.1f}" '
                     f'stroke="{_GRID}" stroke-width="1" stroke-dasharray="4 3"/>')
        parts.append(f'<text x="2" y="{y + 3:.1f}" font-size="8" fill="{_MUTED}">{lbl}</text>')
    if len(pts) >= 2:
        poly = ' '.join(f'{xpos(i):.1f},{ypos(v):.1f}' for i, v in pts)
        parts.append(f'<polyline points="{poly}" fill="none" stroke="{_ACCENT}" stroke-width="2"/>')
    for i, v in pts:
        x, y = xpos(i), ypos(v)
        parts.append(f'<circle cx="{x:.1f}" cy="{y:.1f}" r="3" fill="{_ACCENT}"/>')
        parts.append(f'<text x="{x:.1f}" y="{y - 6:.1f}" font-size="9" font-weight="700" '
                     f'fill="{_ACCENT}" text-anchor="middle">{v:.2f}</text>')
    for i, role in enumerate(roles):
        parts.append(f'<text x="{xpos(i):.1f}" y="{H - pad_b + 16:.1f}" font-size="9" '
                     f'fill="{_MUTED}" text-anchor="middle">{_e(_ROLE_SHORT.get(role, role))}</text>')
    return (f'<svg viewBox="0 0 {W} {H}" style="width:100%;height:auto" '
            f'xmlns="http://www.w3.org/2000/svg">{"".join(parts)}</svg>')


def _chart_ms_variance(msv):
    """Horizontal bars of milestone slip vs baseline (+behind = red, -ahead = green),
    pure HTML so it never clips. Long names ellipsize with a title tooltip."""
    if not msv:
        return '<div class="note">No milestone variance to plot.</div>'
    mx = max((abs(m.get('var') or 0) for m in msv), default=0) or 1
    rows = []
    for m in msv:
        v = m.get('var') or 0
        w = abs(v) / mx * 100
        col = _BAD if v > 0 else (_GOOD if v < 0 else _MUTED)
        name = _e(m.get('name'))
        rows.append(
            f'<div class="mvrow"><div class="mvname" title="{name}">{name}</div>'
            f'<div class="mvbar"><span style="width:{w:.0f}%;background:{col}"></span>'
            f'<b style="color:{col}">{"+" if v > 0 else ""}{v} d</b></div></div>')
    return f'<div class="mvchart">{"".join(rows)}</div>'


def _dashboard(report):
    d = report.get('dashboard', {})
    kpi_cells = []
    for k in d.get('kpis', []):
        if k['value'] is None:
            val = 'n/a'
        elif k['key'] == 'cpli':
            val = f"{k['value']:.2f}"
        else:
            val = k['value']
        unit = ' wd' if k['key'] == 'length' else ''
        delta = '' if k.get('delta') is None else _sd(k['delta'], '')
        # Previous value + variance, so the dashboard reads current vs previous (fix #3).
        pv = k.get('prev')
        if pv is None:
            prev_txt = ''
        else:
            pv = f"{pv:.2f}" if k['key'] == 'cpli' else pv
            prev_txt = f'Previous {pv}{unit}' + (f' · {delta}' if delta else '')
        kpi_cells.append(f'<div class="kpi"><div class="kk">{_e(k["label"])}</div>'
                         f'<div class="kv">{val}{unit} <span class="kcur">current</span></div>'
                         f'<div class="kd">{prev_txt}</div></div>')
    factors = ''.join(f'<li>{_e(f)}</li>' for f in d.get('factors', []))
    dash = (f'<div class="dash"><div class="health"><div class="hgauge">{_cpli(d.get("cpli"))}<span>CPLI</span></div>'
            f'<div><div class="htitle">Critical Path Health</div>'
            f'<div class="hverdict">{_e(d.get("verdict"))}</div><ul class="hfactors">{factors}</ul></div></div>'
            f'<div class="kpis">{"".join(kpi_cells)}</div></div>')

    # Charts — everything on screen must reach the PDF (fix #1).
    charts = d.get('charts', {})
    charts_html = ''
    if charts:
        cn = _chart_crit_near(charts.get('crit_near', {}) or {})
        ct = _chart_cpli(charts.get('cpli_trend', {}) or {})
        mv = _chart_ms_variance(charts.get('ms_variance', []) or [])
        charts_html = (
            '<div class="charts">'
            f'<div class="chartcard"><div class="chartt">Critical &amp; near by schedule</div>{cn}'
            f'<div class="chartlegend"><i style="background:{_BAD}"></i>Critical '
            f'<i style="background:{_WARN}"></i>Near-critical</div></div>'
            f'<div class="chartcard"><div class="chartt">CPLI trend</div>{ct}</div>'
            f'<div class="chartcard"><div class="chartt">Milestone slip vs baseline</div>{mv}</div>'
            '</div>')
    return dash + charts_html


def _lane(lane):
    role = lane.get('role')
    ms = lane.get('milestone') or {}
    if not ms.get('name') and not lane.get('boxes'):
        # This milestone doesn't exist in this schedule — show a placeholder, not an empty box.
        return (f'<div class="lane"><div class="lanehdr"><span class="lanetag lt-{role}">{_e(_ROLE_LABEL.get(role))}</span>'
                f'<span class="lanesub">this milestone is not in this schedule</span></div></div>')
    fin = ms.get('baseline_finish') if role == 'baseline' else ms.get('expected_finish')
    boxes = [f'<div class="msbox"><div class="msflag">◆ Milestone</div><div class="mst">{_e(ms.get("name"))}</div>'
             f'<div class="msr"><span>{"BL Finish" if role == "baseline" else "Exp Finish"}</span><b>{_e(fin)}</b></div>'
             f'<div class="msr"><span>Delay</span><b>{_sd(ms.get("slip_days"))}</b></div></div>']
    lane_boxes = lane.get('boxes', [])
    first_new = next((i for i, x in enumerate(lane_boxes) if x.get('state') == 'new'), -1)
    for i, b in enumerate(lane_boxes):
        st = b.get('state', 'stayed')
        cls = {'new': 'bnew', 'left': 'bleft', 'done': 'bdone'}.get(st, '')
        flag = ('<div class="bflag bfnew">NEW ON PATH</div>' if st == 'new'
                else '<div class="bflag bfleft">LEFT PATH</div>' if st == 'left' else '')
        planned = '—' if b.get('planned') is None else f"{round(b['planned'])}%"
        actual = '—' if b.get('pct') is None else f"{round(b['pct'])}%"
        tf = '—' if b.get('driver_tf') is None else _sd(b.get('driver_tf'), ' wd')
        # A big ▸ leads a box only when it continues the live chain: a stayed/done box, or a NEW box
        # that is not the first. Never before a LEFT box, and never before the START of the new
        # critical path (Ibrahim: "remove the arrow before the new critical path").
        arrow = '' if (st == 'left' or (st == 'new' and i == first_new)) else '<div class="arw">▸</div>'
        boxes.append(
            f'{arrow}<div class="box {cls}">{flag}<div class="bt">{_e(b.get("name"))}</div>'
            f'<div class="bcrumb">{_e(b.get("crumb"))}</div>'
            f'<div class="bgrid">'
            f'<div><div class="bk">Planned</div><div class="bv">{planned}</div></div>'
            f'<div><div class="bk">Actual</div><div class="bv">{actual}</div></div>'
            f'<div><div class="bk">BL finish</div><div class="bv">{_e(b.get("bl_finish"))}</div></div>'
            f'<div><div class="bk">Expected</div><div class="bv">{_e(b.get("exp_finish"))}</div></div>'
            f'<div class="bfull"><div class="bk">Slip / Total float</div><div class="bv">{_sd(b.get("slip_days"))} / {tf}</div></div>'
            f'</div></div>')
    return (f'<div class="lane"><div class="lanehdr"><span class="lanetag lt-{role}">{_e(_ROLE_LABEL.get(role))}</span>'
            f'<span class="lanesub">{_e(lane.get("sub"))}</span></div><div class="chain">{"".join(boxes)}</div></div>')


def _lanes(report, milestone_ids=None):
    blocks = report.get('milestone_paths', [])
    if milestone_ids:                       # limit the PDF to the chosen milestone paths
        wanted = set(milestone_ids)
        blocks = [b for b in blocks if b.get('id') in wanted]
    if not any(any(ln.get('boxes') for ln in b.get('lanes', [])) for b in blocks):
        return '<div class="note">No governing critical path was found in these schedules.</div>'
    legend = (f'<div class="cplegend"><i style="background:{_BAD_BG};border:2px solid {_BAD}"></i>New on critical path '
              f'<i style="background:{_BAD_BG};border:1px dashed {_BAD}"></i>Left the path '
              f'<i style="background:{_BG};border:1px solid {_EDGE}"></i>Stayed '
              f'<i style="background:{_BG};border:1px solid {_GOOD}"></i>Complete</div>')
    out = []
    for b in sorted(blocks, key=lambda x: 0 if x.get('is_governing') else 1):
        flag = '◆ ' if b.get('is_governing') else ''
        hdr = (f'<div class="mphdr"><span class="mpname">{flag}{_e(b.get("name"))}</span>'
               f'<span class="mpfin">Baseline {_e(b.get("baseline_finish")) or "—"} '
               f'· Current {_e(b.get("current_finish")) or "—"} '
               f'· Slip {_sd(b.get("slip_days"))}</span></div>')
        lanes_html = ''.join(_lane(l) for l in b.get('lanes', []))
        out.append(f'<div class="mpblock">{hdr}{lanes_html}</div>')
    return ''.join(out) + legend


def _census(report):
    c = report.get('census', {})
    roles = report.get('roles', [])
    has = lambda r: r in roles
    cur, prev, bl = c.get('current', {}), c.get('previous', {}), c.get('baseline', {})

    def cell(r, txt):
        return f'<td class="num">{txt}</td>' if has(r) else ''

    src = lambda r: (c.get(r, {}) or {}).get('critical_source')
    mark = lambda o: '<sup>†</sup>' if o.get('critical_source') == 'path' else ''

    def cnt(o, n, p, m=False):
        return '—' if o.get(n) is None else f"{o[n]} · <b>{o[p]}%</b>{mark(o) if m else ''}"

    def _red(txt):
        # Every census variance is RED (Ibrahim's rule).
        return f'<td class="num" style="color:{_BAD};font-weight:700">{txt}</td>' if txt is not None else '<td class="num">—</td>'

    def _sum_over(key, rs):
        vals = [c.get(r, {}).get(key) for r in rs if has(r)]
        vals = [v for v in vals if v is not None]
        return round(sum(vals), 1) if vals else None

    def dcell(key, unit, cumulative, cum_roles, a_role, b_role):
        if cumulative:                                     # cumulative sum of the % columns
            s = _sum_over(key, cum_roles)
            return _red(None if s is None else f'+{s}{unit}')
        a, b = c.get(a_role, {}).get(key), c.get(b_role, {}).get(key)   # plain difference
        if a is None or b is None:
            return '<td class="num">—</td>'
        dec = 2 if key == 'cpli' else 1 if key.endswith('_pct') else 0   # % deltas keep one decimal
        dd = round(a - b, dec)
        shown = f'{dd:.{dec}f}' if dec else f'{dd}'
        return _red(f'{"+" if dd > 0 else ""}{shown}{unit}')

    def colh(role, label):
        dd = c.get(role, {}).get('data_date')
        sub = f'<div class="thdd">{_e(dd)}</div>' if dd else ''
        return f'<th class="num">{label}{sub}</th>'

    # (label, getter, delta-key, delta-unit, cumulative)  All rows = PLAIN difference now.
    rows = [
        ('Critical activities (TF ≤ 0)', lambda o: cnt(o, 'critical', 'critical_pct', True), 'critical_pct', ' pts', False),
        ('Near-critical (0 &lt; TF &lt; 10 wd)', lambda o: cnt(o, 'near', 'near_pct'), 'near_pct', ' pts', False),
        ('Critical path length (remaining, wd)', lambda o: '—' if o.get('path_length_wd') is None else f"{o['path_length_wd']} wd", 'path_length_wd', ' wd', False),
        ('Total float · finish (wd)', lambda o: '—' if o.get('total_float_wd') is None else f"{o['total_float_wd']} wd", 'total_float_wd', ' wd', False),
        ('CPLI', lambda o: _cpli(o.get('cpli')), 'cpli', '', False),
    ]
    head = ('<tr><th>Measure</th>'
            + (colh('baseline', 'Baseline') if has('baseline') else '')
            + (colh('previous', 'Previous') if has('previous') else '')
            + (colh('current', 'Current') if has('current') else '')
            + ('<th class="num">Δ period</th>' if has('previous') else '')
            + ('<th class="num">Δ baseline</th>' if has('baseline') else '') + '</tr>')
    body = ''
    for label, get, key, unit, cumulative in rows:
        body += ('<tr><td>' + label + '</td>'
                 + cell('baseline', get(bl)) + cell('previous', get(prev)) + cell('current', get(cur))
                 + (dcell(key, unit, cumulative, ['previous', 'current'], 'current', 'previous') if has('previous') else '')
                 + (dcell(key, unit, cumulative, ['baseline', 'previous', 'current'], 'current', 'baseline') if has('baseline') else '') + '</tr>')
    note = ('<div class="note">Every variance is the <b>plain difference</b> — current minus the compared '
            'schedule (e.g. baseline 28% → current 62% = +34 pts). All shown red.</div>')
    foot = ''
    if any(src(r) == 'path' for r in roles):
        foot = ('<div class="note"><b>†</b> This schedule has no activity float in its export, so its critical '
                'count is the activities on the <b>longest path to completion</b> (not TF ≤ 0) — a different basis '
                'than a floated update.</div>')
    return f'<table class="t"><thead>{head}</thead><tbody>{body}</tbody></table>{note}{foot}'


def _milestones(report):
    rows = report.get('milestones', [])
    roles = report.get('roles', [])
    c = report.get('census', {})
    has = lambda r: r in roles
    if not rows:
        return '<div class="note">No finish milestones found.</div>'

    def slip(v):
        if v is None:
            return '—'
        col = _BAD if v > 0 else _GOOD if v < 0 else _MUTED
        return f'<span style="color:{col};font-weight:700">{"+" if v > 0 else ""}{v} d</span>'

    def colh(role, label):
        dd = c.get(role, {}).get('data_date')
        sub = f'<div class="thdd">{_e(dd)}</div>' if dd else ''
        return f'<th class="num">{label}{sub}</th>'

    head = ('<tr><th>Milestone</th>'
            + (colh('baseline', 'Baseline') if has('baseline') else '')
            + (colh('previous', 'Previous') if has('previous') else '')
            + (colh('current', 'Current') if has('current') else '')
            + '<th class="num">vs Baseline</th>' + ('<th class="num">This period</th>' if has('previous') else '') + '</tr>')
    body = ''
    for m in rows:
        f = m.get('finishes', {})
        body += ('<tr' + (' class="gov"' if m.get('is_governing') else '') + '><td>'
                 + ('◆ ' if m.get('is_governing') else '') + _e(m.get('name')) + '</td>'
                 + (f'<td class="num">{_e(f.get("baseline"))}</td>' if has('baseline') else '')
                 + (f'<td class="num">{_e(f.get("previous"))}</td>' if has('previous') else '')
                 + (f'<td class="num">{_e(f.get("current"))}</td>' if has('current') else '')
                 + f'<td class="num">{slip(m.get("var_vs_baseline_d"))}</td>'
                 + (f'<td class="num">{slip(m.get("var_this_period_d"))}</td>' if has('previous') else '')
                 + '</tr>')
    return f'<table class="t"><thead>{head}</thead><tbody>{body}</tbody></table>'


def _migration(report):
    m = report.get('float_migration')
    if not m:
        return '<div class="note">Load a previous update or a baseline to see how float moved.</div>'
    c = m.get('counts', {})
    base = 'baseline' if report.get('float_migration_base') == 'baseline' else 'last period'
    cards = [('Near → CRITICAL', c.get('near_to_crit', 0), _BAD, 'worsened'),
             ('Safe → near', c.get('safe_to_near', 0), _WARN, 'eroding'),
             ('Critical → recovered', c.get('crit_to_recovered', 0), _GOOD, 'improved'),
             ('Held critical', c.get('held_crit', 0), _MUTED, 'unchanged')]
    cells = ''.join(f'<div class="band"><div class="bandl">{_e(lbl)}</div>'
                    f'<div class="bandv" style="color:{col}">{n}</div>'
                    f'<div class="bandt" style="color:{col}">{tag}</div></div>' for lbl, n, col, tag in cards)
    toward = (c.get('near_to_crit', 0) or 0) + (c.get('safe_to_near', 0) or 0)
    return (f'<div class="bands">{cells}</div>'
            f'<div class="note">{toward} activities lost float toward the critical path vs {base}. Matched by activity ID.</div>')


def _recommendation(report):
    rec = ''.join(f'<li>{_e(r)}</li>' for r in report.get('recommendation', []))
    return (f'<div class="effect">{_e(report.get("effect"))}</div>'
            f'<div class="recoh">Recommendation</div><ol class="reco">{rec}</ol>')


def render_html(report, sections=None, milestone_ids=None, theme='light'):
    secs = [
        ('verdict', '', _banner(report), False),
        ('dashboard', 'Execution dashboard', _dashboard(report), False),
        ('driving_path', 'Critical Path Analyzer', _lanes(report, milestone_ids), True),
        ('census', 'Critical & near-critical census', _census(report), False),
        ('milestones', 'Every milestone · finish comparison & path health', _milestones(report), False),
        ('float_migration', 'Float migration', _migration(report), False),
        ('recommendation', 'Effect on completion & recommendation', _recommendation(report), False),
    ]
    keys = set(sections) if sections else None
    body = [_header(report)]
    for key, title, html, planner in secs:
        if keys is not None and key not in keys:
            continue
        t = f'<h2>{_e(title)}</h2>' if title else ''
        body.append(f'<section data-sec="{key}"{" class=pagebreak" if planner else ""}>{t}{html}</section>')
    return f'''<!doctype html><html><head><meta charset="utf-8"><style>
      @page {{ size: A4 landscape; margin: 11mm; }}
      * {{ box-sizing: border-box; }}
      body {{ font-family: -apple-system, Segoe UI, Roboto, Arial, sans-serif; color: var(--rpt-ink); font-size: 12px; margin: 0; }}
      h1 {{ font-size: 18px; margin: 0; }} h2 {{ font-size: 13px; margin: 16px 0 8px; color: var(--rpt-accent); }}
      section.pagebreak {{ page-break-before: always; }}
      .rh {{ display: flex; justify-content: space-between; align-items: flex-start; border-bottom: 2px solid var(--rpt-edge); padding-bottom: 8px; }}
      .meta {{ color: var(--rpt-muted); font-size: 11px; margin-top: 3px; }}
      .ddstrip {{ display: flex; flex-wrap: wrap; gap: 6px; margin-top: 6px; }}
      .ddpill {{ background: var(--rpt-surface-2); color: var(--rpt-ink-soft); font-size: 10px; border-radius: 5px; padding: 2px 8px; }}
      .ddpill b {{ color: var(--rpt-ink); }}
      .banner {{ display: flex; align-items: center; gap: 11px; border: 1px solid; border-radius: 9px; padding: 10px 14px; margin-top: 10px; }}
      .btag {{ color: var(--rpt-accent-ink); font-size: 10px; font-weight: 800; text-transform: uppercase; letter-spacing: .5px; padding: 3px 9px; border-radius: 5px; }}
      .btext {{ font-size: 13px; font-weight: 600; }}
      .dash {{ display: flex; gap: 12px; }}
      .health {{ flex: 1; display: flex; gap: 12px; align-items: center; border: 1px solid var(--rpt-edge); border-left: 4px solid var(--rpt-bad); border-radius: 10px; padding: 10px 14px; }}
      .hgauge {{ flex: none; width: 74px; height: 74px; border-radius: 50%; border: 6px solid var(--rpt-edge); display: flex; flex-direction: column; align-items: center; justify-content: center; font-size: 19px; font-weight: 800; }}
      .hgauge span {{ font-size: 9px; font-weight: 600; color: var(--rpt-muted); }}
      .htitle {{ font-size: 11px; font-weight: 800; text-transform: uppercase; letter-spacing: .4px; color: var(--rpt-muted); }}
      .hverdict {{ font-size: 12px; font-weight: 700; margin: 3px 0 5px; }}
      .hfactors {{ margin: 0; padding-left: 15px; font-size: 10.5px; color: var(--rpt-muted); line-height: 1.5; }}
      .kpis {{ flex: 1; display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px; }}
      .kpi {{ border: 1px solid var(--rpt-edge); border-radius: 9px; padding: 8px 11px; }}
      .kk {{ font-size: 9.5px; text-transform: uppercase; letter-spacing: .3px; color: var(--rpt-muted); font-weight: 700; }}
      .kv {{ font-size: 19px; font-weight: 800; margin-top: 2px; }}
      .kv .kcur {{ font-size: 9px; font-weight: 700; color: var(--rpt-accent); text-transform: uppercase; letter-spacing: .3px; }}
      .kd {{ font-size: 11px; font-weight: 700; color: var(--rpt-muted); margin-top: 2px; }}
      .charts {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 12px; margin-top: 12px; }}
      .chartcard {{ border: 1px solid var(--rpt-edge); border-radius: 10px; padding: 10px 12px; page-break-inside: avoid; }}
      .chartt {{ font-size: 10.5px; font-weight: 800; text-transform: uppercase; letter-spacing: .3px; color: var(--rpt-muted); margin-bottom: 6px; }}
      .chartlegend {{ font-size: 9.5px; color: var(--rpt-muted); margin-top: 4px; }}
      .chartlegend i {{ display: inline-block; width: 9px; height: 9px; border-radius: 2px; vertical-align: middle; margin: 0 3px 0 8px; }}
      .mvchart {{ display: flex; flex-direction: column; gap: 6px; }}
      .mvrow {{ font-size: 10px; }}
      .mvname {{ color: var(--rpt-ink-soft); white-space: nowrap; overflow: hidden; text-overflow: ellipsis; margin-bottom: 2px; }}
      .mvbar {{ display: flex; align-items: center; gap: 5px; }}
      .mvbar span {{ display: inline-block; height: 10px; border-radius: 3px; min-width: 1px; }}
      .mvbar b {{ font-size: 9.5px; white-space: nowrap; }}
      .mpblock {{ page-break-inside: avoid; margin-bottom: 14px; }}
      .mphdr {{ display: flex; justify-content: space-between; align-items: baseline; gap: 10px; border-bottom: 1px solid var(--rpt-edge); padding-bottom: 4px; margin-bottom: 8px; }}
      .mpname {{ font-size: 12px; font-weight: 800; color: var(--rpt-ink); }}
      .mpfin {{ font-size: 10px; color: var(--rpt-muted); }}
      .lane {{ margin-bottom: 12px; }} .lanehdr {{ display: flex; gap: 9px; align-items: center; margin-bottom: 6px; }}
      .lanetag {{ font-size: 10px; font-weight: 800; text-transform: uppercase; letter-spacing: .5px; padding: 3px 9px; border-radius: 6px; }}
      .lt-baseline {{ background: var(--rpt-surface-2); color: var(--rpt-ink-soft); }} .lt-previous {{ background: var(--rpt-warn-bg); color: var(--rpt-warn); }} .lt-current {{ background: var(--rpt-accent-soft); color: var(--rpt-accent); }}
      .lanesub {{ color: var(--rpt-muted); font-size: 10.5px; }}
      .chain {{ display: flex; align-items: stretch; flex-wrap: wrap; gap: 5px; }}
      .arw {{ display: flex; align-items: center; color: var(--rpt-accent); font-weight: 900; font-size: 26px; padding: 0 1px; }}
      .msbox {{ flex: none; width: 150px; border: 2px solid var(--rpt-accent); border-radius: 10px; padding: 8px 10px; background: var(--rpt-accent-soft); }}
      .msflag {{ font-size: 9px; text-transform: uppercase; color: var(--rpt-accent); margin-bottom: 4px; }}
      .mst {{ font-size: 11.5px; font-weight: 800; color: var(--rpt-accent); margin-bottom: 5px; }}
      .msr {{ display: flex; justify-content: space-between; font-size: 10.5px; margin: 2px 0; }} .msr span {{ color: var(--rpt-muted); }}
      .box {{ flex: none; width: 176px; border: 1px solid var(--rpt-edge); border-radius: 9px; padding: 7px 9px; position: relative; }}
      .bgrid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 3px 8px; margin-top: 2px; }}
      .bgrid .bk {{ font-size: 8px; text-transform: uppercase; }} .bgrid .bv {{ font-size: 10.5px; }}
      .bfull {{ grid-column: 1 / -1; border-top: 1px dashed var(--rpt-edge); padding-top: 3px; margin-top: 1px; }}
      .box.bdone {{ border-color: var(--rpt-good); }}
      .box.bnew {{ border: 2px solid var(--rpt-bad); background: var(--rpt-bad-bg); }}
      .box.bleft {{ border: 1px dashed var(--rpt-hair-strong); background: var(--rpt-surface-2); }}
      .bflag {{ position: absolute; top: -8px; left: 8px; color: var(--rpt-accent-ink); font-size: 8px; font-weight: 800; padding: 1px 6px; border-radius: 4px; }}
      .bfnew {{ background: var(--rpt-bad); }} .bfleft {{ background: var(--rpt-muted); color: var(--rpt-bg); }}
      .bt {{ font-size: 11px; font-weight: 700; }} .bcrumb {{ font-size: 8.5px; color: var(--rpt-muted); margin: 2px 0 6px; }}
      .brow {{ display: flex; justify-content: space-between; font-size: 10px; margin: 2px 0; }}
      .brow.last {{ border-top: 1px dashed var(--rpt-edge); padding-top: 3px; margin-top: 3px; }}
      .bk {{ color: var(--rpt-muted); }} .bv {{ font-weight: 700; }}
      .cplegend {{ font-size: 10px; color: var(--rpt-muted); margin-top: 6px; }}
      .cplegend i {{ display: inline-block; width: 10px; height: 10px; border-radius: 3px; vertical-align: middle; margin: 0 4px 0 12px; }}
      table.t {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
      table.t th, table.t td {{ padding: 6px 9px; text-align: left; border-bottom: 1px solid var(--rpt-edge); }}
      table.t th {{ font-size: 10px; text-transform: uppercase; letter-spacing: .3px; color: var(--rpt-muted); }}
      table.t th .thdd {{ text-transform: none; font-weight: 600; font-size: 9px; color: var(--rpt-muted); margin-top: 2px; }}
      table.t td.num, table.t th.num {{ text-align: right; }}
      table.t tr.gov td {{ font-weight: 800; }}
      .bands {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; }}
      .band {{ border: 1px solid var(--rpt-edge); border-radius: 8px; padding: 9px; text-align: center; }}
      .bandl {{ font-size: 10px; color: var(--rpt-muted); font-weight: 700; }} .bandv {{ font-size: 18px; font-weight: 800; margin-top: 3px; }} .bandt {{ font-size: 10px; font-weight: 700; margin-top: 2px; }}
      .effect {{ font-size: 12px; line-height: 1.5; background: var(--rpt-surface); border-left: 3px solid var(--rpt-accent); border-radius: 0 8px 8px 0; padding: 10px 13px; }}
      .recoh {{ font-size: 11px; font-weight: 800; text-transform: uppercase; letter-spacing: .4px; color: var(--rpt-accent); margin: 12px 0 5px; }}
      .reco {{ margin: 0; padding-left: 18px; font-size: 11.5px; line-height: 1.6; }}
      .note {{ color: var(--rpt-muted); font-size: 11px; margin-top: 8px; }}
    </style>{report_theme.theme_style_tag(theme)}</head><body>{"".join(body)}</body></html>'''


# ── Excel ────────────────────────────────────────────────────────────────────

def to_excel(report, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)
    roles = report.get('roles', [])
    c = report.get('census', {})

    ws = wb.active
    ws.title = 'Census'
    ws.append(['Measure'] + [_ROLE_LABEL[r] for r in roles])
    for cell in ws[1]:
        cell.font = bold

    def crow(label, key, pct=None):
        row = [label]
        for r in roles:
            o = c.get(r, {})
            if pct is None:
                row.append(o.get(key))
            else:
                row.append(f"{o.get(key)} ({o.get(pct)}%)" if o.get(key) is not None else None)
        ws.append(row)

    crow('Total activities', 'total_activities')
    crow('Critical activities', 'critical', 'critical_pct')
    crow('Near-critical', 'near', 'near_pct')
    crow('Critical path length (wd)', 'path_length_wd')
    crow('Total float · finish (wd)', 'total_float_wd')
    crow('CPLI', 'cpli')

    wm = wb.create_sheet('Milestones')
    wm.append(['Milestone', 'Governing'] + [_ROLE_LABEL[r] for r in roles]
              + ['vs Baseline (d)', 'This period (d)', 'CPLI', 'Crit fronts', 'Near fronts'])
    for cell in wm[1]:
        cell.font = bold
    for m in report.get('milestones', []):
        f = m.get('finishes', {})
        wm.append([m.get('name'), 'Yes' if m.get('is_governing') else ''] + [f.get(r) for r in roles]
                  + [m.get('var_vs_baseline_d'), m.get('var_this_period_d'), m.get('cpli'),
                     m.get('crit_fronts'), m.get('near_fronts')])

    wp = wb.create_sheet('Driving path (current)')
    wp.append(['Work front', 'State', 'Actual %', 'Exp finish', 'Slip (d)', 'Float now (wd)'])
    for cell in wp[1]:
        cell.font = bold
    cur_lane = next((l for l in report.get('lanes', []) if l.get('role') == 'current'), None)
    if cur_lane:
        for b in cur_lane.get('boxes', []):
            wp.append([b.get('name'), b.get('state'), round(b.get('pct') or 0), b.get('exp_finish'),
                       b.get('slip_days'), b.get('driver_tf') if b.get('state') == 'left' else None])

    mig = report.get('float_migration')
    if mig:
        wf = wb.create_sheet('Float migration')
        wf.append(['Move', 'Count'])
        wf['A1'].font = bold
        wf['B1'].font = bold
        cc = mig.get('counts', {})
        for lbl, key in [('Near → Critical', 'near_to_crit'), ('Safe → near', 'safe_to_near'),
                         ('Critical → recovered', 'crit_to_recovered'), ('Held critical', 'held_crit')]:
            wf.append([lbl, cc.get(key, 0)])

    wb.save(path)
    return path
