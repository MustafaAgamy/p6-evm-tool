"""Excel I/O for the Productivity KB — all offline (openpyxl), KB-driven.

The template is generated from the LIVE KB (never a fixed list): the Project Types /
Work Types / Methods / Units it offers come from the templates currently loaded, so it
auto-grows as the KB grows — no code change. Nothing here invents a rate; calculation
and the Knowledge-Not-Available decision stay in the engine (calc/match).
"""
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

INPUT_HEADERS = ["Activity ID", "Activity Name", "Project Type",
                 "Type of Work / Discipline", "Quantity (BOQ)", "Unit", "Method"]

_H2K = {
    "activity id": "activity_id", "activity name": "name", "project type": "project_type",
    "type of work / discipline": "work_type", "type of work": "work_type", "discipline": "work_type",
    "quantity (boq)": "quantity", "quantity": "quantity", "qty": "quantity", "boq qty": "quantity",
    "unit": "unit", "method": "method",
}


def classifications(templates):
    """What the current KB can classify — derived live from the loaded templates."""
    ptypes, works, methods, units = set(), set(), set(), set()
    for t in templates:
        if t.get("work_type"):
            works.add(t["work_type"])
        if t.get("method"):
            methods.add(t["method"])
        if t.get("unit"):
            units.add(t["unit"])
        rates = t.get("rates") or ([t["rate"]] if t.get("rate") else [])
        for r in rates:
            pt = (r.get("context") or {}).get("project_type")
            if pt and str(pt).upper() != "GLOBAL":
                ptypes.add(pt)
    return {"project_types": sorted(ptypes), "work_types": sorted(works),
            "methods": sorted(methods), "units": sorted(units)}


def write_template(path, templates):
    cl = classifications(templates)
    wb = Workbook()
    ws = wb.active
    ws.title = "Activities"
    ws.append(INPUT_HEADERS)

    ref = wb.create_sheet("KB Reference")
    ref.append(["Project Types", "Work Types", "Methods", "Units"])
    cols = [cl["project_types"], cl["work_types"], cl["methods"], cl["units"]]
    for i in range(max((len(c) for c in cols), default=0)):
        ref.append([(c[i] if i < len(c) else "") for c in cols])

    # dropdown validations sourced from the reference lists (input col -> reference col)
    for in_col, values, ref_col in [(3, cl["project_types"], 1), (4, cl["work_types"], 2),
                                     (6, cl["units"], 4), (7, cl["methods"], 3)]:
        n = len(values)
        if not n:
            continue
        rc = get_column_letter(ref_col)
        dv = DataValidation(type="list", formula1=f"='KB Reference'!${rc}$2:${rc}${n + 1}",
                            allow_blank=True)
        ws.add_data_validation(dv)
        ic = get_column_letter(in_col)
        dv.add(f"{ic}2:{ic}1000")

    for c in range(1, len(INPUT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 24
    wb.save(path)
    return path


def read_activities(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Activities"] if "Activities" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [_H2K.get(str(h or "").strip().lower()) for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(v not in (None, "") for v in r):
            continue
        a = {}
        for key, val in zip(header, r):
            if key and val not in (None, ""):
                a[key] = val
        if a.get("quantity") is not None:
            try:
                a["quantity"] = float(a["quantity"])
            except (TypeError, ValueError):
                a["quantity"] = None
        if a.get("name"):
            out.append(a)
    return out


MAPPING_HEADERS = ["Activity ID", "Activity Name", "Type of Work", "Project Type", "Method",
                   "Quantity", "Unit", "Productivity Rate", "Rate Unit", "Crew", "Equipment",
                   "Conditions", "Source", "Confidence", "Status"]


def _crew_str(res):
    parts = [f"{c.get('trade')} x{c.get('count')}" for c in (res or {}).get("crew", []) if c.get("trade")]
    if parts:
        return ", ".join(parts)
    mnp = ((res or {}).get("labor") or {}).get("mnp")
    return f"{mnp} persons" if mnp is not None else ""


def write_mapping_sheet(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Productivity Rates"
    ws.append(MAPPING_HEADERS)
    for row in rows:
        inp = row.get("input") or {}
        res = row.get("result")
        rate = (res or {}).get("rate") or {}
        eq = ", ".join((e.get("name") or "") for e in (res or {}).get("equipment", []))
        status = (row.get("status_label") or ("Calculated" if res else "Knowledge Not Available"))
        ws.append([
            inp.get("activity_id", ""), inp.get("name", ""),
            (res or {}).get("work_type") or inp.get("work_type", ""),
            inp.get("project_type", ""), (res or {}).get("method") or inp.get("method", ""),
            inp.get("quantity", ""), inp.get("unit", ""),
            rate.get("value", ""), rate.get("unit", ""),
            _crew_str(res), eq, rate.get("conditions", ""),
            rate.get("source", ""), rate.get("confidence", ""), status,
        ])
    for c in range(1, len(MAPPING_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    wb.save(path)
    return path
